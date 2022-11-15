import pandas as pd
import re
import datetime
import math
import os
import datetime
import glob
from datetime import timedelta
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from openpyxl.styles import Border, Side

from datetime import datetime
start_time = datetime.now()

def octant_analysis(mod=5000):
    # use glob to get all the files
    # in the folder
    path = 'input'
    excel_files = glob.glob(os.path.join(path, "*.xlsx"))
      
    octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}

    # try block for error handling
    try:
        # creating directory if not present
        current_directory = os.getcwd()
        final_directory = os.path.join(current_directory, r'output')
        if not os.path.exists(final_directory):
          os.makedirs(final_directory)
    except:
        print('Error occured while making output folder.')

    # loop over the list of csv files
    for f in excel_files:
        x = f.split('/')
        space_str = ''
        # file name to be used for creating output file
        inp_file_name = x[len(x) -1][:-5]
        # read the excel file
        outp = pd.read_excel(f)
        # stored average of each column in input
        u_avg = outp['U'].mean()
        v_avg = outp['V'].mean()
        w_avg = outp['W'].mean()

        # entered average values in a list
        u_avg_col = [u_avg]
        v_avg_col = [v_avg]
        w_avg_col = [w_avg]

        # filled remaininig cells of average column
        u_avg_col.extend(['']*(len(outp['U'])-1))
        v_avg_col.extend(['']*(len(outp['V'])-1))
        w_avg_col.extend(['']*(len(outp['W'])-1))

        try:
            # entered average column in outp variable
            outp["U Avg"] = u_avg_col
            outp["V Avg"] = v_avg_col
            outp["W Avg"] = w_avg_col
        except:
            print('Error : Mismatch in length of Average quantities columns.')

        u_dash = []
        v_dash = []
        w_dash = []

        # computed other columns
        for i in outp['U']:
            u_dash.append(round(i - u_avg,3))

        for i in outp['V']:
            v_dash.append(round(i - v_avg,3))

        for i in outp['W']:
            w_dash.append(round(i - w_avg,3))

        try:
            # entered the columns in outp variable
            outp["U\'=U - U avg"] = u_dash
            outp["V\'=V - V avg"] = v_dash
            outp["W\'=W - W avg"] = w_dash
        except:
            print('Error : Mismatch in length of colums.')

        # created Octant value column
        octant_col = []
        x = []
        y = []
        z = []

        cnt = len(outp['U'])
        empty_col = []
        empty_col.extend(['']*cnt)

        # filled x,y,z lists
        for i in outp["U\'=U - U avg"]:
            x.append(i)
        for i in outp["V\'=V - V avg"]:
            y.append(i)
        for i in outp["W\'=W - W avg"]:
            z.append(i)

        itr = math.ceil(cnt/mod)

        single_freq = {}

        # computed frequency of each octant in whole range
        for i in range(1, 5):
            single_freq[i] = 0
            single_freq[0-i] = 0

        # finding octant values of each row
        for i in range(cnt):
            cur = i/mod
            cur = cur*mod

            if x[i] >= 0 and y[i] >= 0:  # 1st quadrant
                if z[i] > 0:
                    octant_col.append(1)
                    single_freq[1] += 1
                else:
                    octant_col.append(-1)
                    single_freq[-1] += 1
            elif x[i] <= 0 and y[i] >= 0:  # 2nd quadrant
                if z[i] > 0:
                    octant_col.append(2)
                    single_freq[2] += 1
                else:
                    octant_col.append(-2)
                    single_freq[-2] += 1
            elif x[i] <= 0 and y[i] <= 0:  # 3rd quadrant
                if z[i] > 0:
                    octant_col.append(3)
                    single_freq[3] += 1
                else:
                    octant_col.append(-3)
                    single_freq[-3] += 1
            elif x[i] >= 0 and y[i] <= 0:  # 4th quadrant
                if z[i] > 0:
                    octant_col.append(4)
                    single_freq[4] += 1
                else:
                    octant_col.append(-4)
                    single_freq[-4] += 1

        try:
            outp["Octant"] = octant_col
            space_str = space_str + ' '
            outp[space_str] = empty_col

            mod_col = ['', '', '', 'Mod ' + str(mod)]
            mod_col.extend(['']*(cnt - len(mod_col)))

            space_str = space_str + ' '
            outp[space_str] = mod_col

        except:
            print('Error : Mismatch in length of Octant')

        octant_id = ['', 'Octant ID', 'Overall Count']

        # formed the ranges of Octand_ID column
        for i in range(itr):
            num1 = mod*i
            num2 = (mod*(i+1)) - 1
            if num2 > cnt:
                num2 = cnt - 1

            str_range = str(num1) + '-' + str(num2)
            octant_id.append(str_range)

        octant_id.extend(['']*(len(outp['U']) - len(octant_id)))

        # error handling if column length is mismatched
        try:
            # entered "Octant ID" column
            outp['Overall Octant Count'] = octant_id
        except:
            print('Error encountered : Mismatch in length of columns in Octant ID column')
        
        # iterating for each octant value for range frequency
        # dictionary -> for identification of columns having header as spaces of various sized strings 
        strng_hdr = {}
        for i in range(1, 5):

            # computing for +i octant value
            col_one = ['', i, single_freq[i]]

            # iterating for each range
            for w in range(itr):
                sums = 0
                num1 = w*mod
                num2 = (mod*(w+1)) - 1
                if num2 > cnt:
                    num2 = cnt

                # finding number of elements in each range
                for q in range(cnt):
                    if octant_col[q] == i and num1 <= q and q <= num2:
                        sums += 1

                col_one.append(sums)

            col_one.extend(['']*(len(outp['U']) - len(col_one)))

            # error handling if column length is mismatched
            try:
                # filled +i octant column in outp variable
                space_str = space_str + ' '
                strng_hdr[i] = space_str
                outp[space_str] = col_one
            except:
                print('Error encountered : Mismatch in length of some octant ids column.')

            # computing for -i octant value
            col_one = ['', 0-i, single_freq[0-i]]
            # iterating for each range
            for w in range(itr):
                sums = 0
                num1 = w*mod
                num2 = (mod*(w+1)) - 1
                if num2 > cnt:
                    num2 = cnt
                # finding number of elements in each range
                for q in range(cnt):
                    if octant_col[q] == (0-i) and num1 <= q and q <= num2:
                        sums += 1

                col_one.append(sums)

            # filled the remaining cells of the column
            col_one.extend(['']*(len(outp['U']) - len(col_one)))

            try:
                # filled -i octant column in outp variable
                space_str = space_str + ' '
                strng_hdr[0-i] = space_str
                outp[space_str] = col_one
            except:
                print('Error encountered while entering -ve octant id columns')
        
        # initially filled all cells as empty then filled values in them
        rank_one_octant = ['', 'Rank 1 Octand ID']
        rank_one_octant.extend(['']*(len(outp['U']) - len(rank_one_octant)))

        # iterating for each octant rank finding
        for i in range(1, 5):

            # for +i octant id
            rank_col = ['', 'Rank Octant ' + str(i)]
            
            # iterating for each range for a particular octant id
            for w in range(itr+1):
                ar_cmp = []

                # inserting values in a list to calculate rank
                for b in range(1, 5):
                    ar_cmp.append([outp[strng_hdr[b]][2+w], b])
                    ar_cmp.append([outp[strng_hdr[0-b]][2+w], 0-b])

                # descending order sorting
                ar_cmp.sort(reverse=True) 

                # finding rank
                for b in range(len(ar_cmp)):
                    if ar_cmp[b][1] == i:
                        if b == 0:
                            rank_one_octant[2 + w] = i
                        rank_col.append(b+1)
                        break

            rank_col.extend(['']*(len(outp['U']) - len(rank_col)))

            # try catch statements for error checking
            try:
                space_str = space_str + ' '
                strng_hdr['Rank Octant ' + str(i)] = space_str
                outp[space_str] = rank_col
            except:
                print('Error : Mismatch in length of Rank column during +ve octant entering')

            # proceeding the same as above for -i octant id
            rank_col = ['', 'Rank Octant ' + str(0-i)]
            
            # iterating for each range for a particular octant id
            for w in range(1 + itr):
                ar_cmp = []

                # inserting values in a list to calculate rank
                for b in range(1, 5):
                    ar_cmp.append([outp[strng_hdr[b]][2+w], b])
                    ar_cmp.append([outp[strng_hdr[0-b]][2+w], 0-b])

                ar_cmp.sort(reverse=True)

                # finding rank
                for b in range(len(ar_cmp)):
                    if ar_cmp[b][1] == (0-i):
                        if b == 0:
                            rank_one_octant[2 + w] = (0-i)
                        rank_col.append(b+1)
                        break

            rank_col.extend(['']*(len(outp['U']) - len(rank_col)))

            # try catch statements for error checking
            try:
                space_str = space_str + ' '
                strng_hdr['Rank Octant ' + str(0-i)] = space_str
                outp[space_str] = rank_col
            except:
                print('Error : Mismatch in length of Rank column during +ve octant entering')

        # try catch statements for error avoidance
        try:
            space_str = space_str + ' '
            strng_hdr['Rank1 Octant ID'] = space_str
            outp[space_str] = rank_one_octant
        except:
            print('Error : Mismatch in rank1 octant id column')
        
        # inserting values into list by lookup from given dictionary
        rank_one_name = ['', 'Rank 1 Octant Name']
  
        # outp.to_excel(r'output/'+ inp_file_name + '_octant_analysis_mod_' + str(mod) +'.xlsx', index=0)
        rank_one_name.append(octant_name_id_mapping[str(rank_one_octant[2])])

        for i in range(itr):
            rank_one_name.append(octant_name_id_mapping[str(rank_one_octant[3+i])])

        rank_one_name.extend(['']*(len(outp['U']) - len(rank_one_name)))

        try:
            space_str = space_str + ' '
            strng_hdr['Rank1 Octant Name'] = space_str
            outp[space_str] = rank_one_name
        except:
            print('Error : Mismatch in length of rank1 octant name')

        # entering the lower terms of excel sheet
        cur_cnt = itr+4
        try:
            outp[strng_hdr['Rank Octant 4']][cur_cnt] = 'Octant ID'
            cur_cnt = cur_cnt + 1
        except:
            print('Error : Mismatch in columns lengths')

        for i in range(1, 5):
            try:
                outp[strng_hdr['Rank Octant 4']][cur_cnt] = i
                outp[strng_hdr['Rank Octant 4']][cur_cnt+1] = 0-i
                cur_cnt = cur_cnt + 2
            except:
                print('Error : Mismatch in columns lengths')
        cur_cnt = itr+4
        try:
            outp[strng_hdr['Rank Octant -4']][cur_cnt] = 'Octant Name'
        except:
            print('Error : Mismatch in columns lengths')
        cur_cnt = cur_cnt + 1
        
        # entering names terms of given dictionary
        for i in range(1, 5):
            try:
                outp[strng_hdr['Rank Octant -4']][cur_cnt] = octant_name_id_mapping[str(i)]
                outp[strng_hdr['Rank Octant -4']][cur_cnt+1] = octant_name_id_mapping[str(0-i)]
                cur_cnt = cur_cnt + 2
            except:
                print('Error : Mismath in column lengths')
                
        cur_cnt = itr+4
        try:
            outp[strng_hdr['Rank1 Octant ID']][cur_cnt] = 'Count of Rank 1 Mod Values'
        except:
            print('Error : Mismath in column lengths')
        cur_cnt = cur_cnt + 1
        
        # inserting count of rank1 mod values
        for i in range(1, 5):
            sums = 0
            for w in range(2, 2+itr):
                if outp[strng_hdr['Rank1 Octant ID']][w] == i:
                    sums = sums + 1
            try:
                outp[strng_hdr['Rank1 Octant ID']][cur_cnt] = sums
            except:
                print('Error : Mismath in column lengths')

            sums = 0
            for w in range(2, 2+itr):
                if outp[strng_hdr['Rank1 Octant ID']][w] == (0-i):
                    sums = sums + 1
            try:
                outp[strng_hdr['Rank1 Octant ID']][cur_cnt+1] = sums
            except:
                print('Error : Mismath in column lengths')
            cur_cnt = cur_cnt+2

        space_str = space_str + ' '
        outp[space_str] = empty_col
        from_col = ['','', 'From']
        from_col.extend(['']*13)
        for i in range(itr):
            from_col.append('From')
            from_col.extend(['']*12)

        from_col.extend(['']*(len(outp['U']) - len(from_col)))

        try:
            space_str = space_str + ' '
            outp[space_str] = from_col
        except:
            print('Error : Mismatch in length of columns')

        # formed "Octant ID" column
        octant_id = [' ', 'Octant #']

        # appended +ve and -ve octant values
        for i in range(1,5):
          octant_id.append(i);
          octant_id.append(0 - i);
        
        octant_id.extend(['']*3)

        # entered Mod Transition Count detailed cells
        for i in range(itr):
            octant_id.append('Mod Transition Count')
            num1 = mod*i
            num2 = (mod*(i+1)) - 1
            if num2 > cnt:
                num2 = cnt - 1
                
            str_range = str(num1) + '-' + str(num2)
            octant_id.append(str_range)

            octant_id.append('Octant #')
            for w in range(1,5):
                octant_id.append(w);
                octant_id.append(0 - w);
            octant_id.extend(['']*2)
              

        octant_id.extend(['']*(len(outp['U']) - len(octant_id)))      

        # error handling if column length is mismatched     
        try:
          # entered "Octant ID" column
          outp['Overall Transition Count'] = octant_id
        except:
          print('Error encountered : Mismatch in length of columns in Octant ID column')

        # iterating for each octant value for range frequency
        for i in range(1,5):
            
            # computing for +i octant value
            col_ones = []
            
            if i == 1:
              col_ones.append('To')
            else:
              col_ones.append('')
            col_ones.append(i)

            # created dictionary to store count of octant values in total range
            dictn = {}
            for w in range(1,5):
              dictn[w] = 0
              dictn[0-w] = 0
            
            # incrementing dictionary values when corresponding octant value is found
            for q in range(1, cnt):
              if octant_col[q] == i:
                dictn[octant_col[q - 1]] += 1

            # appending data to dictionary for different octant values transition
            for w in range(1,5):
              col_ones.append(dictn[w])
              col_ones.append(dictn[0-w])
            
            col_ones.append('')

            # iterating over individual ranges 
            for w in range(itr):
              col_ones.extend(['']*3)

              if i == 1:
                col_ones.append('To')
              else:
                col_ones.append('')
              col_ones.append(i)

              num1 = w*mod
              num2 = (mod*(w+1)) - 1
              if num2 > cnt-1:
                num2 = cnt-2

              # created dictionary for individual ranges whose transition is to +i octant value
              indv_dictn = {}
              for q in range(1,5):
                indv_dictn[q] = 0
                indv_dictn[0-q] = 0

              # filled indv_dictn dictionary
              for q in range(num1, num2+1):
                if octant_col[q+1] == i:
                  indv_dictn[octant_col[q]] += 1

              # appending the values of transition of particular octant value
              for q in range(1, 5):
                col_ones.append(indv_dictn[q])
                col_ones.append(indv_dictn[0-q])
            
            # filled the remaining cells of the column
            col_ones.extend(['']*(len(outp['U'])-len(col_ones)))
            # error handling if column length is mismatched          
            try:
              # filled +i octant column in outp variable
              space_str = space_str + ' '
              strng_hdr[str(i)] = space_str
              outp[space_str] = col_ones
            except:
              print('Error encountered : Mismatch in length of some octant ids column.')

            # computing for -i octant value
            col_ones = ['', 0-i]

            # created dictionary to store count of octant values in total range
            dictn = {}
            for w in range(1,5):
              dictn[w] = 0
              dictn[0-w] = 0
            
            # incrementing dictionary values when corresponding octant value is found
            for q in range(1, cnt):
              if octant_col[q] == (0-i):
                dictn[octant_col[q - 1]] += 1

            # appending data to dictionary for different octant values transition
            for w in range(1,5):
              col_ones.append(dictn[w])
              col_ones.append(dictn[0-w])

            # appended with empty cell
            col_ones.append('')
            # iterating over individual ranges for -ve i octant value
            for w in range(itr):
              col_ones.extend(['']*3)
              col_ones.append('')
              col_ones.append(0-i)

              num1 = w*mod
              num2 = (mod*(w+1)) - 1
              if num2 > cnt-1:
                num2 = cnt-2

              # reinitialised dictionary for individual ranges whose transition is to -i octant value
              indv_dictn = {}
              for q in range(1,5):
                indv_dictn[q] = 0
                indv_dictn[0-q] = 0

              # filled indv_dictn dictionary
              for q in range(num1, num2+1):
                if octant_col[q+1] == (0-i):
                  indv_dictn[octant_col[q]] += 1

              # appending the values of transition of particular octant value
              for q in range(1, 5):
                col_ones.append(indv_dictn[q])
                col_ones.append(indv_dictn[0-q])

            # filled the remaining cells of the column
            col_ones.extend(['']*(len(outp['U'])-len(col_ones)))      

            try:
                # filled -i octant column in outp variable
                space_str = space_str + ' '
                strng_hdr[str(0-i)] = space_str
                outp[space_str] = col_ones  
            except:
                print('Error while entering columns')

        space_str = space_str + ' '
        outp[space_str] = empty_col

        long_subs_len = ['', 'Octant ##']
        for i in range(1,5):
            long_subs_len.append(i)
            long_subs_len.append(0-i)
        
        long_subs_len.extend(['']*(len(outp['U']) - len(long_subs_len)))
        outp['Longest Subsequence Length'] = long_subs_len
        # created 2 dictionary
        dictn_len = {}            # dictionary for maximum subsequence length of octant values
        # dictionary for count of maximum subsequence length of octant values
        dictn_cnt = {}

        # initialised dictionary elements with 0
        for i in range(1, 5):
            dictn_len[i] = 0
            dictn_cnt[i] = 0
            dictn_len[0-i] = 0
            dictn_cnt[0-i] = 0

        # updating maximum subsequence length of each octant id in dictn_len
        for w in range(cnt):
            counts = 0                          # variable to store length of current subsequence

            for q in range(w, cnt):
                # if same octant id found incrementing count
                if octant_col[q] == octant_col[w]:
                    counts += 1

                    if q == (cnt - 1):
                        w = q
                else:                                   # else breaking loop and updating subsequence length
                    w = q-1
                    break

            dictn_len[octant_col[w]] = max(counts, dictn_len[octant_col[w]])

        # looping over all octant ids
        for i in range(1, 5):

            # finding maximum subsequence length's count for +i octant value
            # maximum subsequence length = dictn_len[i]

            # iterating for each starting point of subsequence
            for w in range(cnt - dictn_len[i] + 1):
                flag = True

                # checking over the window of length dictn_len[i] starting from w
                for q in range(w, w + dictn_len[i]):
                    if octant_col[q] != i:
                        flag = False
                # updating count if condition satisfied
                if flag == True:
                    dictn_cnt[i] += 1

            # finding maximum subsequence length's count for -i octant value
            # maximum subsequence length = dictn_len[0-i]

            # iterating for each starting point of subsequence
            for w in range(cnt - dictn_len[0-i] + 1):
                flag = True

                # checking over the window of length dictn_len[i] starting from w
                for q in range(w, w + dictn_len[0-i]):
                    if octant_col[q] != (0-i):
                        flag = False

                # updating count if condition satisfied
                if flag == True:
                    dictn_cnt[0-i] += 1

        # list for longest subsequence length and appending the values respectivelu by lookup from dictionary
        longest_len = ['', 'Longest Subsequence Length']
        for i in range(1, 5):
            longest_len.append(dictn_len[i])
            longest_len.append(dictn_len[0-i])

        longest_len.extend(['']*(len(outp['U']) - len(longest_len)))

        try:
            # inserted the column into inp
            space_str = space_str + ' '
            outp[space_str] = longest_len
        except:
            print('Error : Mismatch in length of longest subsequence length columns')

        # list for longest subsequence length count
        longest_count = ['', 'Count']
        for i in range(1, 5):
            longest_count.append(dictn_cnt[i])
            longest_count.append(dictn_cnt[0-i])

        longest_count.extend(['']*(len(outp['U']) - len(longest_count)))
        try:
            # inserted the column into inp
            space_str = space_str + ' '
            outp[space_str] = longest_count

            space_str = space_str + ' '
            outp[space_str] = empty_col
        except:
            print('Error : Mismatch in length of count of subsequences')

        count_col = ['', 'Octant ###']
        for i in range(1, 5):
            # for +i
            count_col.append(i)
            count_col.append('Time')
            count_col.extend(['']*(dictn_cnt[i]))

            # for -i
            count_col.append(0-i)
            count_col.append('Time')
            count_col.extend(['']*(dictn_cnt[0-i]))

        count_col.extend(['']*(len(outp['U']) - len(count_col)))

        try:
            outp['Longest Subsequence Length with Range'] = count_col
        except:
            print('Error : Mismatch in length of columns named Count.')

        long_len = ['', 'Longest Subsequence Length']

        # looping over all octant ids
        for i in range(1, 5):

            # for +i desired operations
            long_len.append(dictn_len[i])
            long_len.append('From')

            # maximum subsequence length = dictn_len[i]
            # iterating for each starting point of subsequence
            for w in range(cnt - dictn_len[i] + 1):
                flag = True

                # checking over the window of length dictn_len[i] starting from w
                for q in range(w, w + dictn_len[i]):
                    if octant_col[q] != i:
                        flag = False
                # updating count if condition satisfied

                if flag == True:
                    long_len.append(outp['T'][w])

            # for -i desired operations
            long_len.append(dictn_len[0-i])
            long_len.append('From')

            # maximum subsequence length = dictn_len[0-i]
            # iterating for each starting point of subsequence
            for w in range(cnt - dictn_len[0-i] + 1):
                flag = True

                # checking over the window of length dictn_len[i] starting from w
                for q in range(w, w + dictn_len[0-i]):
                    if octant_col[q] != (0-i):
                        flag = False
                # updating count if condition satisfied
                if flag == True:
                    long_len.append(outp['T'][w])

        long_len.extend(['']*(len(outp['U']) - len(long_len)))

        # checking for errors in mismatch of lengths
        try:
            space_str = space_str + ' '
            outp[space_str] = long_len
        except:
            print('Error : Mismatch in length of column named Longest Subsequence Length')

        long_cnt = ['', 'Count']
        # looping over all octant ids
        for i in range(1, 5):

            # for +i desired operations
            long_cnt.append(dictn_cnt[i])
            long_cnt.append('To')

            # maximum subsequence length = dictn_len[i]
            # iterating for each starting point of subsequence
            for w in range(cnt - dictn_len[i] + 1):
                flag = True

                # checking over the window of length dictn_len[i] starting from w
                for q in range(w, w + dictn_len[i]):
                    if octant_col[q] != i:
                        flag = False
                # updating count if condition satisfied
                if flag == True:
                    long_cnt.append(outp['T'][w + dictn_len[i] - 1])

            # for -i desired operations
            long_cnt.append(dictn_cnt[0-i])
            long_cnt.append('To')

            # maximum subsequence length = dictn_len[0-i]
            # iterating for each starting point of subsequence
            for w in range(cnt - dictn_len[0-i] + 1):
                flag = True

                # checking over the window of length dictn_len[i] starting from w
                for q in range(w, w + dictn_len[0-i]):
                    if octant_col[q] != (0-i):
                        flag = False
                # updating count if condition satisfied
                if flag == True:
                    long_cnt.append(outp['T'][w + dictn_len[0-i] - 1])

        long_cnt.extend(['']*(len(outp['U']) - len(long_cnt)))

        # checking for errors in mismatch of lengths
        try:
            space_str = space_str + ' '
            strng_hdr['long_cnt'] = space_str
            outp[space_str] = long_cnt
        except:
            print('Error : Mismatch in length of another column named Count.')

        # list of all the cells which are to be colored
        be_filled = []
        # entering all the required cells to be colored
        col_dictn = {1:'W', -1:'X', 2:'Y', -2:'Z', 3:'AA', -3:'AB', 4:'AC', -4:'AD'}
        for w in range(2, 3 + itr):
            be_filled.append(col_dictn[outp[strng_hdr['Rank1 Octant ID']][w]] + str(w+2))
        
        for w in range(len(outp['U'])):
            temp_flag = bool(0)
            for ep in range(1,5):
                if outp['Overall Transition Count'][w] == ep or outp['Overall Transition Count'][w] == (0-ep):
                    temp_flag = 1
                    break

            if temp_flag == 1:
                temp_cmp_list = []
                alp_cur = ord('J')
                for ep in range(1,5):
                    temp_cmp_list.append([outp[strng_hdr[str(ep)]][w], 'A' + chr(alp_cur) + str(w+2)])
                    temp_cmp_list.append([outp[strng_hdr[str(0-ep)]][w], 'A' + chr(alp_cur + 1) + str(w+2)])
                    alp_cur = alp_cur + 2

                temp_cmp_list.sort(reverse = True)
                be_filled.append(temp_cmp_list[0][1])

        thin = Side(border_style="thin", color='000000')
        # list of all cells having border
        final_list = []
        border_list = ['N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF']

        # entering all cells required with a border
        len_list = len(border_list)
        for i in range(len_list):
            for w in range(3, itr + 5):
                final_list.append(border_list[i] + str(w))
        
        border_list = ['AC', 'AD', 'AE']
        for i in range(len(border_list)):
            for w in range(itr+6, itr + 15):
                final_list.append(border_list[i] + str(w))

        border_list = ['AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ']
        for i in range(len(border_list)):
            for w in range(3, 12):
                final_list.append(border_list[i] + str(w))        

            for q in range(0, itr):
                temp_cell = 17 + (q*13)
                for w in range(9):
                    final_list.append(border_list[i] + str(w + temp_cell))

        border_list = ['AS', 'AT', 'AU']
        for i in range(len(border_list)):
            for w in range(3, 12):
                final_list.append(border_list[i] + str(w))

        border_list = ['AW', 'AX', 'AY']
        row_start = 1
        while 1:
            if outp[strng_hdr['long_cnt']][row_start] == '':
                break
            final_list.append('AW' + str(row_start + 2))
            final_list.append('AX' + str(row_start + 2))
            final_list.append('AY' + str(row_start + 2))
            row_start = row_start + 1

        # making workbook object
        wb = Workbook()
        ws = wb.active
        # making worksheet dataframe
        for r in dataframe_to_rows(outp, index=False, header=True):
            ws.append(r)

        # colour pattern for filling in cells
        fill_cell1 = PatternFill(patternType='solid', fgColor='FFFF00')
        # filling colours
        for w in be_filled:
            ws[w].fill = fill_cell1
        
        # drawing borders
        for w in final_list:
            ws[w].border = Border(top=thin, left=thin, right=thin, bottom=thin)

        wb.save(r'output/'+ inp_file_name + '_octant_analysis_mod_' + str(mod) +'.xlsx')
             
mod=5000
octant_analysis(5000)
